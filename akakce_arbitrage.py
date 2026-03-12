#!/usr/bin/env python3
"""
akakce_arbitrage.py
───────────────────
Scans Akakçe "Fark Atan Fiyatlar" and reports electronics products where:
  • The cheapest seller is Amazon / Amazon Prime / MediaMarkt
  • The price gap to the 2nd cheapest offer is > MIN_PRICE_GAP_PCT

Results are saved to:
  • PostgreSQL  (arbitrage_runs + arbitrage_items tables)
  • Excel file  (local runs only — skipped when OUTPUT_DIR is None)

INSTALL (once):
    pip install playwright openpyxl psycopg2-binary python-dotenv
    playwright install chromium

RUN:
    python akakce_arbitrage.py
"""

import asyncio
import os
import re
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from playwright.async_api import async_playwright, Page, Browser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION  —  adjust these without touching logic
# ══════════════════════════════════════════════════════════════════════════════

BASE_URL          = "https://www.akakce.com/fark-atan-fiyatlar/"
MAX_PAGES         = 8      # how many listing pages to scan
PAGE_DELAY_MS     = 2000   # wait after each page load (ms)
PRODUCT_DELAY_MS  = 1500   # wait between product pages (ms)
MIN_PRICE_GAP_PCT = 5.0    # minimum % gap to include a product
OUTPUT_DIR        = Path(".")  # where Excel files are saved

# Seller names exactly as Akakçe renders them in img[alt] attributes.
# Key  = friendly label used in the output.
# List = every img[alt] value that maps to that label.
TARGET_SELLERS: dict[str, list[str]] = {
    "Amazon":       ["Amazon Türkiye"],
    "Amazon Prime": ["Amazon Prime"],
    "MediaMarkt":   ["Media Markt"],
}

# Flat set for O(1) membership test
_TARGET_ALTS: set[str] = {v for vals in TARGET_SELLERS.values() for v in vals}


# ══════════════════════════════════════════════════════════════════════════════
# SELECTORS  —  update here if Akakçe changes its HTML structure
# ══════════════════════════════════════════════════════════════════════════════

SEL = {
    # Regex that identifies a product-detail URL
    "product_url_re": re.compile(r"akakce\.com/[^?#]+,\d+\.html"),

    # Product page: each seller is one <li>
    # Confirmed structure: ul.pl_v9 > li  (fallback: ul[class*="pl_v"] > li)
    "seller_rows_js": 'ul.pl_v9 > li, ul[class*="pl_v"] > li',

    # Within a row: big-retailer logo → img[alt] = seller name
    "seller_logo_js": "span.v_v8 img[alt]",

    # Within a row: small-retailer text node
    "seller_text_js": "span.v_v8",

    # Within a row: campaign / discounted price (preferred)
    "price_campaign_js": "span.pt_v8.cmpgn_pt_v8",

    # Within a row: regular price (excludes crossed-out orig price)
    "price_regular_js": "span.pt_v8:not(.orig_pt_v8):not(.cmpgn_pt_v8)",

    # Within a row: offer / buy link
    "offer_link_js": "a[href]",

    # Electronics breadcrumb detection: anchor whose text is exactly this
    "elec_anchor_text": "Elektronik",
}


# ══════════════════════════════════════════════════════════════════════════════
# LOGGING
# ══════════════════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("akakce_arb")


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def parse_price(text: str) -> Optional[float]:
    """
    Parse Turkish price string → float.
    '12.499,00 TL' → 12499.0
    '1.234,56'     → 1234.56
    Returns None if parsing fails.
    """
    s = re.sub(r"\s+", "", text or "")
    # Main pattern: digits (with . thousands sep), comma, 2 decimal digits
    m = re.search(r"([\d.]+),(\d{2})", s)
    if m:
        try:
            return float(m.group(1).replace(".", "") + "." + m.group(2))
        except ValueError:
            pass
    # Fallback: plain integer-like number
    m2 = re.search(r"[\d.]+", s)
    if m2:
        try:
            return float(m2.group(0).replace(".", ""))
        except ValueError:
            pass
    return None


def friendly_label(raw: str) -> str:
    """Map an Akakçe img[alt] value to the friendly seller label."""
    for label, alts in TARGET_SELLERS.items():
        if raw.strip() in alts:
            return label
    return raw.strip()


def is_target(raw: str) -> bool:
    """True if this raw seller name is one of our buy-source targets."""
    return raw.strip() in _TARGET_ALTS


# ══════════════════════════════════════════════════════════════════════════════
# BROWSER SETUP
# ══════════════════════════════════════════════════════════════════════════════

async def new_page(browser: Browser) -> Page:
    """Create a browser context + page with Turkish headers."""
    ctx = await browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        locale="tr-TR",
        extra_http_headers={"Accept-Language": "tr-TR,tr;q=0.9"},
    )
    return await ctx.new_page()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Collect product links from the listing pages
# ══════════════════════════════════════════════════════════════════════════════

async def collect_links(page: Page) -> list[str]:
    """
    Paginate through BASE_URL and collect unique product-detail URLs.
    Stops early when a page adds zero new links.
    """
    seen: set[str] = set()

    for p in range(1, MAX_PAGES + 1):
        url = BASE_URL if p == 1 else f"{BASE_URL}?p={p}"
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=30_000)
            await page.wait_for_timeout(PAGE_DELAY_MS)

            all_hrefs: list[str] = await page.evaluate(
                "() => Array.from(document.querySelectorAll('a[href]')).map(a => a.href)"
            )

            # Keep only product-detail URLs, deduplicate within page
            links = list(dict.fromkeys(
                h for h in all_hrefs if SEL["product_url_re"].search(h)
            ))

            new_count = sum(1 for l in links if l not in seen)
            seen.update(links)
            log.info(f"  Listing p{p}: {len(links)} links (+{new_count} new, total {len(seen)})")

            if new_count == 0:
                log.info("  No new links on this page — stopping.")
                break

        except Exception as e:
            log.error(f"  Listing p{p} error: {e}")
            break

    return list(seen)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Scrape a single product page
# ══════════════════════════════════════════════════════════════════════════════

# JavaScript evaluated inside the browser — keeps all DOM work client-side.
# Note: Python's triple-quoted string passes this verbatim to Playwright.
#       Inside JS regex literals, \s and \/ are standard JS escapes.
_SCRAPE_JS = """
() => {
    // ── Electronics detection ──────────────────────────────────────────────
    // Akakçe breadcrumb uses plain <a> tags. If any anchor's exact text is
    // "Elektronik" the product is in the electronics category.
    const isElectronics = Array.from(document.querySelectorAll('a'))
        .some(a => a.textContent.trim() === 'Elektronik');

    // ── Product name ───────────────────────────────────────────────────────
    const h1 = document.querySelector('h1[class*="v_h"], h1.v_h, h1');
    const name = h1 ? h1.textContent.trim().replace(/\\s+/g, ' ') : '';

    // ── Seller rows ────────────────────────────────────────────────────────
    const rows = document.querySelectorAll('ul.pl_v9 > li, ul[class*="pl_v"] > li');
    const sellers = [];

    rows.forEach(li => {
        // Big retailers (Amazon, MediaMarkt) render a logo: <img alt="Amazon Türkiye">
        // Small retailers render plain text inside span.v_v8
        const span = li.querySelector('span.v_v8');
        let sellerRaw = '';
        if (span) {
            const img = span.querySelector('img[alt]');
            sellerRaw = img
                ? img.alt.trim()
                : span.textContent.trim().replace(/^\\//, '').replace(/\\s+/g, ' ');
        }

        // Price: prefer campaign/discounted span, fall back to regular price span
        // Crossed-out original prices (.orig_pt_v8) are intentionally excluded
        const priceEl =
            li.querySelector('span.pt_v8.cmpgn_pt_v8') ||
            li.querySelector('span.pt_v8:not(.orig_pt_v8):not(.cmpgn_pt_v8)');
        const priceRaw = priceEl ? priceEl.textContent : '';

        // Offer / buy link for this seller row
        const a = li.querySelector('a[href]');
        const offerUrl = a ? a.href : '';

        if (sellerRaw && priceRaw) {
            sellers.push({ sellerRaw, priceRaw, offerUrl });
        }
    });

    return { isElectronics, name, sellers };
}
"""


async def scrape_product(page: Page, url: str) -> Optional[dict]:
    """
    Load a product page and extract sellers + prices.

    Returns:
        {
            is_electronics: bool,
            name: str,
            sellers: [{seller_raw, seller_label, price, offer_url}, ...]
        }
        or None on repeated failure.
    """
    # Up to 3 attempts to load the page
    for attempt in range(1, 4):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=30_000)
            await page.wait_for_timeout(PAGE_DELAY_MS)
            break
        except Exception as e:
            log.warning(f"    [attempt {attempt}/3] load failed — {e}")
            if attempt == 3:
                return None
            await asyncio.sleep(3)

    try:
        raw = await page.evaluate(_SCRAPE_JS)
    except Exception as e:
        log.error(f"    JS evaluate error: {e}")
        return None

    # Parse prices in Python (cleaner than doing it in JS)
    parsed_sellers = []
    for s in raw.get("sellers", []):
        price = parse_price(s["priceRaw"])
        if price and price > 0:
            parsed_sellers.append({
                "seller_raw":   s["sellerRaw"],
                "seller_label": friendly_label(s["sellerRaw"]),
                "price":        price,
                "offer_url":    s.get("offerUrl", ""),
            })

    return {
        "is_electronics": raw.get("isElectronics", False),
        "name":           raw.get("name", ""),
        "sellers":        parsed_sellers,
    }


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Business logic: detect arbitrage opportunity
# ══════════════════════════════════════════════════════════════════════════════

def find_arbitrage(sellers: list[dict]) -> Optional[dict]:
    """
    Returns an arbitrage dict if ALL conditions are met:
      1. At least 2 sellers with valid prices
      2. The cheapest seller is Amazon / Amazon Prime / MediaMarkt
      3. (2nd price − cheapest) / cheapest  >  MIN_PRICE_GAP_PCT %

    Returns None otherwise.
    """
    if len(sellers) < 2:
        return None

    by_price = sorted(sellers, key=lambda s: s["price"])
    cheapest = by_price[0]
    second   = by_price[1]

    if not is_target(cheapest["seller_raw"]):
        return None

    gap_tl  = second["price"] - cheapest["price"]
    gap_pct = (gap_tl / cheapest["price"]) * 100

    if gap_pct <= MIN_PRICE_GAP_PCT:
        return None

    return {
        "cheapest": cheapest,
        "second":   second,
        "gap_tl":   round(gap_tl, 2),
        "gap_pct":  round(gap_pct, 2),
    }


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — Export results to Excel
# ══════════════════════════════════════════════════════════════════════════════

# (column_header, column_width)
_COLUMNS = [
    ("run_timestamp",             20),
    ("product_name",              55),
    ("product_page_url",          62),
    ("cheapest_seller",           18),
    ("cheapest_price",            16),
    ("cheapest_offer_url",        62),
    ("second_cheapest_seller",    22),
    ("second_cheapest_price",     18),
    ("second_cheapest_offer_url", 62),
    ("price_difference_tl",       22),
    ("price_difference_percent",  24),
]


def export_excel(results: list[dict], run_ts: str) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arbitrage"

    # ── Header row ────────────────────────────────────────────────────────
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, (name, width) in enumerate(_COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    # ── Data rows (sorted by gap % descending) ────────────────────────────
    for row_i, r in enumerate(
        sorted(results, key=lambda r: r["arb"]["gap_pct"], reverse=True), 2
    ):
        arb = r["arb"]
        row_values = [
            run_ts,
            r["name"],
            r["url"],
            arb["cheapest"]["seller_label"],
            arb["cheapest"]["price"],
            arb["cheapest"]["offer_url"],
            arb["second"]["seller_label"],
            arb["second"]["price"],
            arb["second"]["offer_url"],
            arb["gap_tl"],
            arb["gap_pct"],
        ]
        for col, val in enumerate(row_values, 1):
            ws.cell(row=row_i, column=col, value=val)

        # Green highlight for especially high-gap rows (> 15%)
        if arb["gap_pct"] > 15:
            green = PatternFill("solid", fgColor="E2EFDA")
            for col in range(1, len(_COLUMNS) + 1):
                ws.cell(row=row_i, column=col).fill = green

    timestamped = OUTPUT_DIR / f"arbitrage_report_{run_ts}.xlsx"
    latest      = OUTPUT_DIR / "latest_report.xlsx"
    wb.save(timestamped)
    wb.save(latest)
    return timestamped


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 — Save results to PostgreSQL
# ══════════════════════════════════════════════════════════════════════════════

def db_connect():
    """Connect to PostgreSQL using DATABASE_URL from environment."""
    url = os.environ.get("DATABASE_URL")
    if not url:
        log.warning("DATABASE_URL not set — skipping DB save.")
        return None
    try:
        conn = psycopg2.connect(url, sslmode="require")
        return conn
    except Exception as e:
        log.error(f"DB connect error: {e}")
        return None


def db_ensure_tables(conn):
    """Create arbitrage tables if they don't exist (idempotent)."""
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS arbitrage_runs (
                id            SERIAL PRIMARY KEY,
                run_timestamp TEXT NOT NULL,
                total_scanned INTEGER DEFAULT 0,
                total_results INTEGER DEFAULT 0,
                created_at    TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS arbitrage_items (
                id                   SERIAL PRIMARY KEY,
                run_id               INTEGER NOT NULL REFERENCES arbitrage_runs(id) ON DELETE CASCADE,
                product_name         TEXT,
                product_url          TEXT,
                cheapest_seller      TEXT,
                cheapest_price       REAL,
                cheapest_offer_url   TEXT,
                second_seller        TEXT,
                second_price         REAL,
                second_offer_url     TEXT,
                gap_tl               REAL,
                gap_pct              REAL,
                created_at           TIMESTAMP DEFAULT NOW()
            )
        """)
    conn.commit()


def db_save(conn, results: list[dict], run_ts: str, total_scanned: int):
    """Insert a run record + all result rows into PostgreSQL."""
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO arbitrage_runs (run_timestamp, total_scanned, total_results) "
                "VALUES (%s, %s, %s) RETURNING id",
                (run_ts, total_scanned, len(results))
            )
            run_id = cur.fetchone()[0]

            rows = [
                (
                    run_id,
                    r["name"],
                    r["url"],
                    r["arb"]["cheapest"]["seller_label"],
                    r["arb"]["cheapest"]["price"],
                    r["arb"]["cheapest"]["offer_url"],
                    r["arb"]["second"]["seller_label"],
                    r["arb"]["second"]["price"],
                    r["arb"]["second"]["offer_url"],
                    r["arb"]["gap_tl"],
                    r["arb"]["gap_pct"],
                )
                for r in results
            ]
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO arbitrage_items
                   (run_id, product_name, product_url,
                    cheapest_seller, cheapest_price, cheapest_offer_url,
                    second_seller, second_price, second_offer_url,
                    gap_tl, gap_pct)
                   VALUES %s""",
                rows,
            )
        conn.commit()
        log.info(f"DB: {len(results)} sonuç kaydedildi (run_id={run_id})")
    except Exception as e:
        log.error(f"DB save error: {e}")
        conn.rollback()


# ══════════════════════════════════════════════════════════════════════════════
# TERMINAL SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def print_summary(results: list[dict], total_scanned: int, stats: dict):
    sep = "─" * 70
    print(f"\n{sep}")
    print("  AKAKÇE ARBİTRAJ TARAMA SONUCU")
    print(sep)
    print(f"  Taranan ürün              : {total_scanned}")
    print(f"  Elektronik                : {stats['elec']}")
    print(f"  En ucuz = hedef satıcı    : {stats['is_target']}")
    print(f"  Fark > {MIN_PRICE_GAP_PCT}% → SONUÇ      : {len(results)}")
    print(f"  Atlandı (elektronik ×)    : {stats['not_elec']}")
    print(f"  Atlandı (hedef satıcı ×)  : {stats['not_target']}")
    print(f"  Atlandı (fark yetersiz)   : {stats['gap_low']}")
    print(f"  Hata                      : {stats['errors']}")
    print(sep)

    if not results:
        print("  Uygun fırsat bulunamadı.\n")
        return

    print(f"\n  EN İYİ FIRSATLAR (% farka göre sıralı):\n")
    for i, r in enumerate(
        sorted(results, key=lambda r: r["arb"]["gap_pct"], reverse=True)[:10], 1
    ):
        a = r["arb"]
        print(f"  {i:2}. {r['name'][:60]}")
        print(f"      Siz alırsınız : {a['cheapest']['seller_label']:<16} "
              f"@ {a['cheapest']['price']:>12,.2f} ₺")
        print(f"      Siz satarsınız: {a['second']['seller_label']:<16} "
              f"@ {a['second']['price']:>12,.2f} ₺")
        print(f"      Fark          : {a['gap_tl']:,.2f} ₺  ({a['gap_pct']:.1f}%)")
        print()


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

async def main():
    run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    stats: dict[str, int] = {
        "elec": 0, "is_target": 0,
        "not_elec": 0, "not_target": 0, "gap_low": 0, "errors": 0,
    }
    results: list[dict] = []
    seen_urls: set[str] = set()

    log.info(f"Akakçe arbitraj taraması başlıyor  [run={run_ts}]")

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage"],
        )
        page = await new_page(browser)

        # ── 1. Collect product links ───────────────────────────────────────
        log.info("\n▶ Ürün linkleri toplanıyor…")
        links = await collect_links(page)
        total = len(links)
        log.info(f"\n  {total} benzersiz ürün linki bulundu.\n")

        if not links:
            log.warning("Hiç link bulunamadı. Çıkılıyor.")
            await browser.close()
            return

        # ── 2+3. Scrape each product and apply filters ─────────────────────
        for i, url in enumerate(links, 1):
            if url in seen_urls:   # safety dedup
                continue
            seen_urls.add(url)
            prefix = f"[{i:3}/{total}]"

            product = await scrape_product(page, url)

            if product is None:
                stats["errors"] += 1
                continue

            if not product["is_electronics"]:
                stats["not_elec"] += 1
                log.debug(f"{prefix} skip — elektronik değil")
                continue

            stats["elec"] += 1

            arb = find_arbitrage(product["sellers"])

            if arb is None:
                # Classify the skip reason for stats
                sellers_sorted = sorted(product["sellers"], key=lambda s: s["price"])
                if len(sellers_sorted) < 2 or not is_target(sellers_sorted[0]["seller_raw"]):
                    stats["not_target"] += 1
                else:
                    stats["gap_low"] += 1
                log.debug(f"{prefix} skip — fırsat yok")
                continue

            stats["is_target"] += 1
            results.append({"name": product["name"], "url": url, "arb": arb})

            log.info(
                f"{prefix} ✓  {product['name'][:38]:38}  "
                f"{arb['cheapest']['seller_label']} {arb['cheapest']['price']:,.0f}₺ → "
                f"{arb['second']['seller_label']} {arb['second']['price']:,.0f}₺  "
                f"(+{arb['gap_pct']:.1f}%)"
            )

            await asyncio.sleep(PRODUCT_DELAY_MS / 1000)

        await browser.close()

    # ── 4. Export to Excel (local runs) ───────────────────────────────────
    if results:
        try:
            out = export_excel(results, run_ts)
            log.info(f"\nExcel kaydedildi  : {out.resolve()}")
            log.info(f"Son rapor         : {(OUTPUT_DIR / 'latest_report.xlsx').resolve()}")
        except Exception as e:
            log.warning(f"Excel kaydedilemedi (normal on server): {e}")

    # ── 5. Save to PostgreSQL ──────────────────────────────────────────────
    conn = db_connect()
    if conn:
        db_ensure_tables(conn)
        db_save(conn, results, run_ts, total)
        conn.close()

    print_summary(results, total, stats)


if __name__ == "__main__":
    asyncio.run(main())
